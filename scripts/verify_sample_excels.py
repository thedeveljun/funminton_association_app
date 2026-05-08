# -*- coding: utf-8 -*-
"""
업로드 파서(lib/screens/clubs/upload_screen.dart) 동작을 모사해서
생성된 엑셀 파일이 검증을 통과하는지 확인.
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = ROOT / "assets" / "excel"


def normalize(s: str) -> str:
    return s.replace(" ", "").replace("*", "").strip()


def parse(path: Path, kind: str) -> tuple[list[dict], list[str], dict]:
    target_sheet = "클럽 임원 등록" if kind == "officer" else "클럽 회원 등록"
    keyword_alt = "임원" if kind == "officer" else "회원"
    keywords = (
        ["클럽명", "회장", "총무"]
        if kind == "officer"
        else ["이름", "성별", "급수", "클럽"]
    )
    core_keyword = "클럽명" if kind == "officer" else "이름"

    wb = load_workbook(path, data_only=True)

    # 시트 탐색 — 파서와 동일하게 dict 순회 (insertion order)
    sheet = None
    for name in wb.sheetnames:
        if target_sheet in name or keyword_alt in name:
            sheet = wb[name]
            break
    if sheet is None:
        for name in wb.sheetnames:
            if wb[name].max_row > 2:
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))

    # 헤더 행 자동 탐지 (첫 5행, 키워드 매칭 최대)
    header_idx = 1
    max_match = 0
    for i in range(min(5, len(rows))):
        matched = 0
        for cell in rows[i]:
            v = "" if cell is None else str(cell)
            for kw in keywords:
                if kw in v:
                    matched += 1
                    break
        if matched > max_match:
            max_match = matched
            header_idx = i

    headers = [
        ("" if c is None else str(c).strip()) for c in rows[header_idx]
    ]

    info = {
        "selected_sheet": sheet.title,
        "header_row": header_idx + 1,  # 1-based for display
        "headers": headers,
    }

    # 핵심 컬럼 인덱스
    core_idx = -1
    for j, h in enumerate(headers):
        if core_keyword in normalize(h):
            core_idx = j
            break

    parsed: list[dict] = []
    errors: list[str] = []

    for i in range(header_idx + 1, len(rows)):
        raw = rows[i]
        rec = {}
        is_example = False
        for j in range(min(len(raw), len(headers))):
            v = "" if raw[j] is None else str(raw[j]).strip()
            if headers[j]:
                rec[headers[j]] = v
                if j == 0 and v == "예시":
                    is_example = True
        if is_example:
            continue
        if core_idx >= 0:
            v = "" if (core_idx >= len(raw) or raw[core_idx] is None) else str(raw[core_idx]).strip()
            if not v:
                continue
        parsed.append(rec)

    def gv(row: dict, *cands: str) -> str:
        for k in row:
            nk = normalize(k)
            for c in cands:
                if nk == normalize(c):
                    return row[k]
        return ""

    for i, row in enumerate(parsed):
        if kind == "officer":
            if not gv(row, "클럽명"):
                errors.append(f"{i+1}번째: 클럽명 필수")
            if not gv(row, "회장이름", "회장 이름"):
                errors.append(f"{i+1}번째: 회장이름 필수")
        else:
            if not gv(row, "클럽명", "소속클럽명"):
                errors.append(f"{i+1}번째: 클럽명 필수")
            if not gv(row, "이름"):
                errors.append(f"{i+1}번째: 이름 필수")
            g = gv(row, "성별")
            if not g:
                errors.append(f"{i+1}번째: 성별 필수")
            elif g not in ("남", "여"):
                errors.append(f"{i+1}번째: 성별 남/여")
            grade = gv(row, "급수")
            if not grade:
                errors.append(f"{i+1}번째: 급수 필수")
            elif grade not in ("A조", "B조", "C조", "D조", "초심조"):
                errors.append(f"{i+1}번째: 급수 오류")
            bd = gv(row, "생년월일")
            if bd:
                digits = re.sub(r"\D", "", bd)
                if len(digits) not in (6, 8):
                    errors.append(f"{i+1}번째: 생년월일 6/8자리")

    return parsed, errors, info


def report(name: str, path: Path, kind: str) -> int:
    print(f"\n=== {name} ===")
    rows, errs, info = parse(path, kind)
    print(f"  선택된 시트:  {info['selected_sheet']}")
    print(f"  헤더 행 위치: {info['header_row']}행")
    print(f"  감지된 헤더:  {info['headers']}")
    print(f"  파싱 데이터:  {len(rows)}건  (예시·빈행 제외)")
    print(f"  검증 오류:   {len(errs)}건")
    for e in errs[:10]:
        print(f"    - {e}")
    return 0 if not errs else 1


def main() -> None:
    rc = 0
    rc |= report("클럽_회원_샘플.xlsx", ASSET_DIR / "클럽_회원_샘플.xlsx", "player")
    rc |= report("클럽_임원_샘플.xlsx", ASSET_DIR / "클럽_임원_샘플.xlsx", "officer")
    if rc:
        print("\n[FAIL]")
        sys.exit(1)
    print("\n[OK] 두 양식 모두 빈 상태에서 오류 없이 파싱됨.")


if __name__ == "__main__":
    main()
