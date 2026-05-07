# -*- coding: utf-8 -*-
"""
클럽_회원_샘플.xlsx / 클럽_임원_샘플.xlsx 재생성 스크립트.

업로드 파서(lib/screens/clubs/upload_screen.dart)와 정확히 호환되도록 만든다.

검증된 패턴 (이전 정상 동작 양식 + 사용자 요청 컬럼명):
- 시트 1: '작성 안내'   — 양식 사용법 (파서가 무시; 행 수 ≤ 2)
- 시트 2: '클럽 회원 등록' / '클럽 임원 등록'  — 실제 입력
    row 1: 제목 (B1:H1 머지, A1 비움 → 헤더 자동탐지에 영향 없음)
    row 2: 헤더 (사용자 요청 그대로: No, 클럽명, 이름, 성별, …)
    row 3: 예시 1행 (1열 = '예시' → 파서가 자동 제외)
    row 4+: 입력용 빈 행 (No 열만 미리 채움 → 핵심 컬럼이 비어 자동 제외)

요청 컬럼:
- 회원: No | 클럽명 | 이름 | 성별 | 급수 | 생년월일 | 전화번호 | 주소
- 임원: No | 클럽명 | 회장이름 | 회장연락처 | 총무이름 | 총무연락처 | 운동장소 | 운동요일
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = PROJECT_ROOT / "assets" / "excel"

# 색상
INK = "FF0D1B3E"
BLUE = "FF2563EB"
SOFT = "FFF4F6FA"
HEADER_BG = "FFEFF6FF"
EXAMPLE_BG = "FFFFF7ED"
GUIDE_BG = "FFFEF9C3"
BORDER_GREY = "FFCBD5E1"

THIN = Side(border_style="thin", color=BORDER_GREY)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    headers: list[str],
    example: list,
    widths: list[int],
    blank_rows: int,
) -> None:
    ws = wb.create_sheet(sheet_name)
    cols = len(headers)

    # ── row 1: 제목 (B1:H1 머지 — A1은 비워두어 헤더 자동탐지를 방해하지 않게) ──
    ws.cell(row=1, column=2, value=title)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=2)
    cell.font = Font(name="맑은 고딕", size=14, bold=True, color=INK)
    cell.alignment = CENTER
    cell.fill = PatternFill("solid", fgColor=SOFT)
    ws.row_dimensions[1].height = 30

    # ── row 2: 헤더 (사용자 요청 그대로) ──
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(name="맑은 고딕", size=11, bold=True, color=BLUE)
        c.alignment = CENTER
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.border = BORDER_ALL
    ws.row_dimensions[2].height = 26

    # ── row 3: 예시 한 행 (1열 = '예시' → 파서가 스킵) ──
    for i in range(1, cols + 1):
        v = example[i - 1] if i - 1 < len(example) else ""
        c = ws.cell(row=3, column=i, value=v)
        c.font = Font(name="맑은 고딕", size=10, italic=True, color="FFB45309")
        c.alignment = CENTER if i == 1 else LEFT
        c.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
        c.border = BORDER_ALL

    # ── row 4+: 빈 입력행 (No 열만 1, 2, 3, … 미리 채움) ──
    for k in range(blank_rows):
        r = 4 + k
        for i in range(1, cols + 1):
            c = ws.cell(row=r, column=i)
            c.alignment = CENTER if i == 1 else LEFT
            c.border = BORDER_ALL
            if i == 1:
                c.value = k + 1
                c.font = Font(name="맑은 고딕", size=10, color="FF94A3B8")

    # 열 너비 + 1행 고정
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def write_guide_sheet(
    wb: Workbook, sheet_name: str, lines: list[tuple[str, str]]
) -> None:
    """양식 안내 (별도 시트). 파서는 시트명에 '회원'/'임원'이 들어간 시트를 우선 선택하므로
    안내 시트명에는 그 키워드를 넣지 않는다."""
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    ws.cell(row=1, column=1, value="양식 작성 안내")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    c = ws.cell(row=1, column=1)
    c.font = Font(name="맑은 고딕", size=14, bold=True, color=INK)
    c.alignment = CENTER
    c.fill = PatternFill("solid", fgColor=SOFT)
    ws.row_dimensions[1].height = 30

    for idx, (label, text) in enumerate(lines, start=2):
        a = ws.cell(row=idx, column=1, value=label)
        a.font = Font(name="맑은 고딕", size=11, bold=True, color=BLUE)
        a.alignment = Alignment(horizontal="right", vertical="center")
        a.fill = PatternFill("solid", fgColor=GUIDE_BG)

        b = ws.cell(row=idx, column=2, value=text)
        b.font = Font(name="맑은 고딕", size=11, color=INK)
        b.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, indent=1
        )
        ws.row_dimensions[idx].height = 22


def build_member_workbook(out: Path) -> None:
    headers = [
        "No",
        "클럽명",
        "이름",
        "성별",
        "급수",
        "생년월일",
        "전화번호",
        "주소",
    ]
    widths = [6, 18, 12, 8, 10, 14, 18, 36]
    example = [
        "예시",
        "과천클럽",
        "홍길동",
        "남",
        "C조",
        "850315",
        "010-1234-5678",
        "경기도 과천시 별양로 12",
    ]

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    write_guide_sheet(
        wb,
        "작성 안내",
        [
            ("업로드 순서", "1) 먼저 ‘클럽_임원_샘플.xlsx’로 클럽을 등록한 뒤  2) 본 파일로 회원 업로드"),
            ("필수 항목", "클럽명, 이름, 성별, 급수  (생년월일은 입력 시 형식 검증)"),
            ("성별", "‘남’ 또는 ‘여’ 만 입력"),
            ("급수", "A조 / B조 / C조 / D조 / 초심조  (정확히 한 글자+조)"),
            ("생년월일", "6자리(예 850315) 또는 8자리(예 19850315)"),
            ("클럽명", "사전에 ‘클럽관리’에 등록된 이름과 정확히 일치해야 함"),
            ("예시 행", "3행은 예시이며 업로드 시 자동 제외됩니다"),
            ("빈 행", "데이터가 없는 행(클럽명/이름 비어있음)은 자동 무시"),
        ],
    )

    write_data_sheet(
        wb,
        sheet_name="클럽 회원 등록",
        title="클럽 회원 등록",
        headers=headers,
        example=example,
        widths=widths,
        blank_rows=50,
    )

    # 데이터 시트가 첫 번째로 보이도록 순서 조정
    wb.move_sheet("클럽 회원 등록", offset=-1)

    wb.save(out)


def build_officer_workbook(out: Path) -> None:
    headers = [
        "No",
        "클럽명",
        "회장이름",
        "회장연락처",
        "총무이름",
        "총무연락처",
        "운동장소",
        "운동요일",
    ]
    widths = [6, 18, 12, 16, 12, 16, 24, 16]
    example = [
        "예시",
        "과천클럽",
        "김회장",
        "010-1111-2222",
        "박총무",
        "010-3333-4444",
        "과천시민회관 체육관",
        "화/목",
    ]

    wb = Workbook()
    wb.remove(wb.active)

    write_guide_sheet(
        wb,
        "작성 안내",
        [
            ("필수 항목", "클럽명, 회장이름"),
            ("운동요일 예", "월/수/금, 화·목, 매일 등 자유 입력"),
            ("중복 처리", "동일 ‘클럽명’이 이미 등록되어 있으면 업로드 시 자동 건너뜀"),
            ("예시 행", "3행은 예시이며 업로드 시 자동 제외됩니다"),
            ("빈 행", "클럽명이 비어있는 행은 자동 무시"),
            ("다음 단계", "클럽 등록 후 ‘클럽_회원_샘플.xlsx’로 회원 업로드"),
        ],
    )

    write_data_sheet(
        wb,
        sheet_name="클럽 임원 등록",
        title="클럽 임원 등록",
        headers=headers,
        example=example,
        widths=widths,
        blank_rows=30,
    )

    wb.move_sheet("클럽 임원 등록", offset=-1)

    wb.save(out)


def build_entry_workbook(out: Path) -> None:
    """대회 참가신청 샘플 — 클럽이 회원에게 받아 협회에 제출하는 양식."""
    headers = [
        "No",
        "이름",
        "소속클럽",
        "종목",
        "급수",
        "파트너이름",
        "파트너소속클럽",
        "연락처",
    ]
    widths = [6, 12, 18, 10, 10, 12, 18, 16]
    example = [
        "예시",
        "홍길동",
        "과천클럽",
        "혼복",
        "C조",
        "김영희",
        "안양클럽",
        "010-1234-5678",
    ]

    wb = Workbook()
    wb.remove(wb.active)

    write_guide_sheet(
        wb,
        "작성 안내",
        [
            ("작성 주체", "각 클럽 총무가 회원에게 신청을 받아 본 양식을 작성"),
            ("필수 항목", "이름, 소속클럽 (둘 다 비어있으면 자동 무시)"),
            ("종목", "혼복 / 남복 / 여복 / 단식  (정확히 입력)"),
            ("파트너이름", "복식(혼복/남복/여복) 신청 시 함께 출전할 파트너 이름. 단식은 비워둠."),
            (
                "파트너소속클럽",
                "파트너가 다른 클럽일 수 있으므로 소속을 함께 기재. 같은 클럽이면 동일하게 입력.",
            ),
            (
                "파트너 페어",
                "두 사람 모두(같은 엑셀 또는 다른 클럽 엑셀) 같은 종목·서로의 이름·서로의 클럽으로 등록되어 있어야 정상 매칭",
            ),
            ("급수", "A조 / B조 / C조 / D조 / 초심조 / S조 / 자강조  (A 만 입력해도 자동 정규화)"),
            (
                "중복",
                "같은 사람·같은 종목이 여러 번 들어오면 마지막 값으로 덮어씁니다",
            ),
            (
                "여러 클럽 누적",
                "각 클럽이 따로 작성한 파일을 순서대로 업로드하면 자동으로 합쳐집니다",
            ),
            ("미등록 회원", "선수관리에 없는 사람은 업로드 시 자동 신규 등록됩니다"),
            ("예시 행", "3행은 예시이며 업로드 시 자동 제외됩니다"),
        ],
    )

    write_data_sheet(
        wb,
        sheet_name="대회 참가신청",
        title="대회 참가신청",
        headers=headers,
        example=example,
        widths=widths,
        blank_rows=80,
    )

    wb.move_sheet("대회 참가신청", offset=-1)
    wb.save(out)


def main() -> None:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)

    member_path = ASSET_DIR / "클럽_회원_샘플.xlsx"
    officer_path = ASSET_DIR / "클럽_임원_샘플.xlsx"
    entry_path = ASSET_DIR / "대회_참가신청_샘플.xlsx"

    build_member_workbook(member_path)
    build_officer_workbook(officer_path)
    build_entry_workbook(entry_path)

    print(f"[OK] {member_path}")
    print(f"[OK] {officer_path}")
    print(f"[OK] {entry_path}")


if __name__ == "__main__":
    main()
